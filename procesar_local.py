"""
=============================================================================
PROCESADOR DE CARGABILIDAD - CNEL EP
Dirección de Operaciones · Gerencia de Distribución
=============================================================================

USO:
  1. Pon los archivos Excel de cada mes en su carpeta correspondiente:
     C:\\Users\\pedro.ulloa\\Desktop\\Operaciones 2025\\Proyectos Cargabilidad\\Cargabilidad 2026\\Enero 2026\\
     C:\\Users\\pedro.ulloa\\Desktop\\Operaciones 2025\\Proyectos Cargabilidad\\Cargabilidad 2026\\Febrero 2026\\
     etc.

  2. Abre CMD y ejecuta:
     python procesar_local.py

  3. Se generará el archivo: datos.json
     Cópialo a la carpeta del repositorio GitHub y haz Push.

=============================================================================
"""

import os
import glob
import json
import subprocess
import tempfile
import shutil
import warnings
warnings.filterwarnings('ignore')

from openpyxl import load_workbook
import pandas as pd

# ─── CONFIGURACIÓN DE RUTAS ────────────────────────────────────────────────────
BASE_2026 = r"C:\Users\pedro.ulloa\Desktop\Operaciones 2025\Proyectos Cargabilidad\Cargabilidad 2026"
BASE_2025 = r"C:\Users\pedro.ulloa\Desktop\Operaciones 2025\Proyectos Cargabilidad\Cargabilidad 2025"

SALIDA_HTML     = r"C:\Users\pedro.ulloa\Documents\GitHub\cargabilidad-cnel\index.html"
PLANTILLA_HTML  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'plantilla.html')

# ─── MAPEO DE MESES ─────────────────────────────────────────────────────────────
MESES = {
    'Enero':      ('2026-01', BASE_2026),
    'Febrero':    ('2026-02', BASE_2026),
    'Marzo':      ('2026-03', BASE_2026),
    'Abril':      ('2026-04', BASE_2026),
    'Mayo':       ('2026-05', BASE_2026),
    'Junio':      ('2026-06', BASE_2026),
    'Julio':      ('2026-07', BASE_2026),
    'Agosto':     ('2026-08', BASE_2026),
    'Septiembre': ('2026-09', BASE_2026),
    'Octubre':    ('2026-10', BASE_2026),
    'Noviembre':  ('2026-11', BASE_2026),
    'Diciembre':  ('2026-12', BASE_2026),
}

MESES_LABELS = {
    '2026-01': 'Enero 2026',   '2026-02': 'Febrero 2026',  '2026-03': 'Marzo 2026',
    '2026-04': 'Abril 2026',   '2026-05': 'Mayo 2026',     '2026-06': 'Junio 2026',
    '2026-07': 'Julio 2026',   '2026-08': 'Agosto 2026',   '2026-09': 'Septiembre 2026',
    '2026-10': 'Octubre 2026', '2026-11': 'Noviembre 2026','2026-12': 'Diciembre 2026',
    '2025-01': 'Enero 2025',   '2025-02': 'Febrero 2025',  '2025-03': 'Marzo 2025',
    '2025-04': 'Abril 2025',   '2025-05': 'Mayo 2025',     '2025-06': 'Junio 2025',
    '2025-07': 'Julio 2025',   '2025-08': 'Agosto 2025',   '2025-09': 'Septiembre 2025',
    '2025-10': 'Octubre 2025', '2025-11': 'Noviembre 2025','2025-12': 'Diciembre 2025',
}

MESES_LABELS_CORTO = {k: v.replace(' 20', " '") for k, v in MESES_LABELS.items()}

# ─── MAPEO DE HOJAS POR UN ─────────────────────────────────────────────────────
# Qué hoja leer según el código de UN
HOJAS_UN = {
    'BOL': ['CARGABILIDAD'],
    'EOR': ['ENERO 2026','FEBRERO 2026','MARZO 2026','ABRIL 2026','MAYO 2026',
            'JUNIO 2026','JULIO 2026','AGOSTO 2026','SEPTIEMBRE 2026','OCTUBRE 2026',
            'NOVIEMBRE 2026','DICIEMBRE 2026'],
    'ESM': ['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO',
            'JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE'],
    'GLR': ['TOTAL'],
    'GYE': ['DEMANDA'],
    'LRS': ['DICIEMBRE 2025','ENERO 2026','FEBRERO 2026','MARZO 2026'],  # LRS usa nombre anterior
    'MAN': ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
            'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'],
    'MLG': ['ENERO 2026','FEBRERO 2026','MARZO 2026','ABRIL 2026','MAYO 2026',
            'JUNIO 2026','JULIO 2026','AGOSTO 2026','SEPTIEMBRE 2026','OCTUBRE 2026',
            'NOVIEMBRE 2026','DICIEMBRE 2026'],
    'STD': ['FORMATO ACTUAL'],
    'STE': ['ENERO 2026','FEBRERO 2026','MARZO 2026','ABRIL 2026','MAYO 2026',
            'JUNIO 2026','JULIO 2026','AGOSTO 2026','SEPTIEMBRE 2026','OCTUBRE 2026',
            'NOVIEMBRE 2026','DICIEMBRE 2026'],
    'SUC': ['Cargabilidad Trasformadores'],
}

# Palabras clave en el nombre del archivo para detectar la UN
PALABRAS_UN = {
    'BOL': ['BOL','BOLIV'],
    'EOR': ['EOR','EL ORO','ELORO'],
    'ESM': ['ESM','ESMER'],
    'GLR': ['GLR','GUAYAS'],
    'GYE': ['GYE','GUAYAQUIL'],
    'LRS': ['LRS','LOS RIOS','LOSRIOS'],
    'MAN': ['MAN','MANABI','MANABÍ'],
    'MLG': ['MLG','MILAGRO'],
    'STD': ['STD','SANTO DOMINGO','SANTODOMINGO'],
    'STE': ['STE','SANTA ELENA','SANTAELENA'],
    'SUC': ['SUC','SUCUMB'],
}

# ─── SUBESTACIONES PARTICULARES A EXCLUIR ──────────────────────────────────────
PARTICULARES = {
    'MAN': ['EL CAFÉ','MOLINERA','LA FABRIL','REFINERÍA DEL PACIFICO','MALL PACÍFICO',
            'EL CEIBAL','CRM','SESME','AVEPECHICHAL','MONTECRISTI GOLF CLUB','FADESA'],
    'EOR': ['PASEO SHOPPING','AUTORIDAD PORTUARIA','INCARPALM','LA GRAN PIAZZA',
            'GOLDEN VALLEY','PETROECUADOR','ENERJUBONES'],
}

# ─── CORRECCIONES MANUALES ─────────────────────────────────────────────────────
REASIGNACIONES_TRAFO = {
    ('MAN','MANTA 3','LA PRADERA'): 'T2',
    ('MAN','MANTA 3','SAN PEDRO'):  'T2',
    ('MAN','MANTA 4','BARBASQUILLO'): 'T2',
    ('STE','SALINAS','*'): 'T1',
}

UN_ORDEN  = ['BOL','EOR','ESM','GLR','GYE','LRS','MAN','MLG','STD','STE','SUC']
UN_NOMBRES = {
    'BOL':'Bolívar','EOR':'El Oro','ESM':'Esmeraldas','GLR':'Guayas-Los Ríos',
    'GYE':'Guayaquil','LRS':'Los Ríos','MAN':'Manabí','MLG':'Milagro',
    'STD':'Santo Domingo','STE':'Santa Elena','SUC':'Sucumbíos'
}

# ─── FUNCIONES AUXILIARES ──────────────────────────────────────────────────────
def detectar_un(nombre_archivo):
    """Detecta el código de UN a partir del nombre del archivo."""
    nombre = nombre_archivo.upper()
    for un, palabras in PALABRAS_UN.items():
        for p in palabras:
            if p in nombre:
                return un
    return None

def hoja_tiene_formato_correcto(ws):
    """
    Verifica si una hoja tiene el formato estándar de cargabilidad.
    Busca palabras clave en las primeras filas: ALIMENTADOR, TRANSFORMADOR, CARGABILIDAD.
    """
    try:
        # Recopilar texto de las primeras 9 filas (encabezados)
        texto = ''
        max_col = min(ws.max_column, 20)
        for r in range(1, 10):
            for c in range(1, max_col+1):
                v = ws.cell(row=r, column=c).value
                if v:
                    texto += str(v).upper() + ' '
        
        # La hoja correcta debe tener TODOS estos términos en las cabeceras
        palabras_clave = ['ALIMENTADOR', 'TRANSFORMADOR', 'CARGABILIDAD']
        return all(p in texto for p in palabras_clave)
    except:
        return False


def detectar_hoja(wb, un, periodo=None):
    """Detecta la hoja correcta. Prioridad:
    1. Hoja con formato correcto que contenga el nombre del mes/año
    2. Hoja con formato correcto (cualquiera)
    3. Hojas conocidas por UN (fallback)
    """
    hojas_disponibles = wb.sheetnames
    
    # Filtrar solo hojas con el formato correcto
    hojas_validas = []
    for h in hojas_disponibles:
        try:
            if hoja_tiene_formato_correcto(wb[h]):
                hojas_validas.append(h)
        except:
            continue
    
    # Si no hay hojas válidas, usar todas
    if not hojas_validas:
        hojas_validas = hojas_disponibles
    
    MESES_NUM = {1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',
                 7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}
    
    if periodo:
        anio, mes_num = periodo.split('-')
        mes_num = int(mes_num)
        nombre_mes = MESES_NUM[mes_num]
        
        # 1. Buscar entre hojas válidas una que contenga el nombre del mes y año
        for h in hojas_validas:
            h_upper = h.upper()
            if nombre_mes in h_upper and anio in h:
                return h
        
        # 2. Buscar hoja válida que contenga solo el nombre del mes
        for h in hojas_validas:
            if nombre_mes in h.upper():
                return h
    
    # 3. Si solo hay UNA hoja válida, usar esa (caso típico de hoja única con datos)
    if len(hojas_validas) == 1:
        return hojas_validas[0]
    
    # 4. Buscar hojas conocidas por UN entre las válidas
    candidatas = HOJAS_UN.get(un, [])
    for h in candidatas:
        if h in hojas_validas:
            return h
    
    # 5. Última opción: primera hoja válida
    return hojas_validas[0] if hojas_validas else hojas_disponibles[0]

def expandir_combinadas(ws):
    """Expande celdas combinadas del worksheet."""
    valores = {}
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            valores[(cell.row, cell.column)] = cell.value
    for rng in ws.merged_cells.ranges:
        val = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row+1):
            for c in range(rng.min_col, rng.max_col+1):
                valores[(r, c)] = val
    return valores

def normalizar_trafo(val):
    """Normaliza el campo TRAFO."""
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('','NAN'): return None
    if s.replace('.','').isdigit():
        n = int(float(s))
        return 'T1' if n == 0 else f'T{n}'
    return s

def convertir_xls_a_xlsx(ruta_xls):
    """
    Convierte un .xls a .xlsx preservando celdas combinadas.
    Usa xlrd para leer incluyendo merged cells y openpyxl para escribir.
    """
    try:
        import xlrd
        wb_xls = xlrd.open_workbook(ruta_xls, formatting_info=True)
        tmp = tempfile.mktemp(suffix='.xlsx')
        
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        wb_new = Workbook()
        wb_new.remove(wb_new.active)  # quitar hoja por defecto
        
        for sheet_name in wb_xls.sheet_names():
            ws_xls = wb_xls.sheet_by_name(sheet_name)
            ws_new = wb_new.create_sheet(title=sheet_name)
            
            # Primero escribir todos los valores
            for row in range(ws_xls.nrows):
                for col in range(ws_xls.ncols):
                    cell = ws_xls.cell(row, col)
                    val = cell.value
                    # Convertir tipo xlrd a Python
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            import datetime
                            val = xlrd.xldate_as_datetime(val, wb_xls.datemode)
                        except:
                            pass
                    elif cell.ctype == xlrd.XL_CELL_EMPTY:
                        val = None
                    ws_new.cell(row=row+1, column=col+1, value=val)
            
            # Luego replicar celdas combinadas
            for crange in ws_xls.merged_cells:
                rlo, rhi, clo, chi = crange
                val = ws_xls.cell(rlo, clo).value
                # Escribir el valor en todas las celdas del rango
                for r in range(rlo, rhi):
                    for c in range(clo, chi):
                        ws_new.cell(row=r+1, column=c+1, value=val)
        
        wb_new.save(tmp)
        return tmp
        
    except Exception as e:
        print(f"    ⚠ Error convirtiendo {os.path.basename(ruta_xls)}: {e}")
        # Fallback: conversión simple sin merged cells
        try:
            df_temp = pd.read_excel(ruta_xls, engine='xlrd', sheet_name=None, header=None)
            tmp = tempfile.mktemp(suffix='.xlsx')
            with pd.ExcelWriter(tmp, engine='openpyxl') as writer:
                for sn, df in df_temp.items():
                    df.to_excel(writer, sheet_name=sn, index=False, header=False)
            return tmp
        except:
            return None

def procesar_archivo(ruta, un, periodo):
    """Lee un archivo Excel y devuelve DataFrame de alimentadores."""
    # Convertir .xls si es necesario
    tmp_file = None
    if ruta.lower().endswith('.xls'):
        tmp_file = convertir_xls_a_xlsx(ruta)
        if not tmp_file:
            return pd.DataFrame()
        ruta_leer = tmp_file
    else:
        ruta_leer = ruta

    try:
        wb = load_workbook(ruta_leer, data_only=True)
        hoja = detectar_hoja(wb, un, periodo)
        ws = wb[hoja]
        # Mostrar qué hoja se está leyendo
        print(f"        [hoja: {hoja}]")
        valores = expandir_combinadas(ws)
    except Exception as e:
        print(f"    ⚠ Error leyendo {os.path.basename(ruta)}: {e}")
        if tmp_file and os.path.exists(tmp_file):
            os.remove(tmp_file)
        return pd.DataFrame()

    # Detectar fila de corte (sección PARTICULARES en GYE)
    fila_corte = ws.max_row + 1
    for r in range(10, ws.max_row+1):
        for c in range(1, 6):
            v = valores.get((r, c))
            if v and 'PARTICULAR' in str(v).upper():
                fila_corte = r
                break
        if fila_corte != ws.max_row + 1:
            break

    rows = []
    for r in range(10, fila_corte):
        sub      = valores.get((r, 3))
        cod_alim = valores.get((r, 4))
        nom_alim = valores.get((r, 5))

        if not sub or not (cod_alim or nom_alim): continue
        sub_str   = str(sub).strip()
        sub_upper = sub_str.upper()
        if 'TOTAL' in sub_upper: continue
        if len(sub_str) > 60: continue
        if any(p in sub_upper for p in ['OBSERVACION','NOTA:','NOTAS:','PERTENEC',
                                         'TRANSFERID','INGRESA A OPE','INDISPONI']): continue

        rows.append({
            'PERIODO': periodo, 'UN_CODIGO': un, 'SUBESTACION': sub_str,
            'CODIGO_ALIM': cod_alim, 'NOMBRE_ALIM': nom_alim,
            'DEM_MIN_MW':      valores.get((r, 6)),
            'DEM_MEDIA_MW':    valores.get((r, 7)),
            'DEM_MAX_MW':      valores.get((r, 8)),
            'FP_ALIM':         valores.get((r, 9)),
            'TRAFO_RAW':       valores.get((r,10)),
            'MVA_OA':          valores.get((r,11)),
            'MVA_FOA':         valores.get((r,12)),
            'DEM_MAX_REG_MES': valores.get((r,13)),
            'FP_TRAFO':        valores.get((r,14)),
            'CARG_OA':         valores.get((r,15)),
            'CARG_FOA':        valores.get((r,16)),
        })

    df = pd.DataFrame(rows)
    if tmp_file and os.path.exists(tmp_file):
        os.remove(tmp_file)
    if len(df) == 0: return df

    df['TRAFO'] = df['TRAFO_RAW'].apply(normalizar_trafo)

    # Reasignaciones manuales
    for (un_r, sub_r, nom_r), trafo_c in REASIGNACIONES_TRAFO.items():
        if un_r != un: continue
        mask_sub = df['SUBESTACION'].str.upper() == sub_r.upper()
        if nom_r == '*':
            df.loc[mask_sub, 'TRAFO'] = trafo_c
        else:
            mask = mask_sub & df['NOMBRE_ALIM'].astype(str).str.upper().str.startswith(nom_r.upper())
            df.loc[mask, 'TRAFO'] = trafo_c

    # Corrección MANTA 3
    if un == 'MAN':
        manta3 = df[df['SUBESTACION'].str.upper() == 'MANTA 3']
        idx_p = manta3[manta3['NOMBRE_ALIM'].astype(str).str.upper().str.startswith('LA PRADERA')].index
        if len(idx_p) > 0:
            mask = (df['SUBESTACION'].str.upper() == 'MANTA 3') & (df.index >= idx_p[0])
            df.loc[mask, 'TRAFO'] = 'T2'

    # Convertir numéricos
    for c in ['DEM_MIN_MW','DEM_MEDIA_MW','DEM_MAX_MW','FP_ALIM',
              'MVA_OA','MVA_FOA','DEM_MAX_REG_MES','FP_TRAFO','CARG_OA','CARG_FOA']:
        df[c] = pd.to_numeric(df[c], errors='coerce')

    # Filtrar particulares
    if un in PARTICULARES:
        df = df[~df['SUBESTACION'].str.upper().isin([s.upper() for s in PARTICULARES[un]])]

    return df.reset_index(drop=True)

def procesar_carpeta_mes(carpeta_mes, periodo):
    """Procesa todos los Excel de una carpeta de mes."""
    archivos = glob.glob(os.path.join(carpeta_mes, '*.xlsx')) + \
               glob.glob(os.path.join(carpeta_mes, '*.xls'))

    encontrados = {}
    for ruta in archivos:
        nombre = os.path.basename(ruta)
        un = detectar_un(nombre)
        if un:
            encontrados[un] = ruta
        else:
            print(f"    ⚠ No se pudo detectar UN para: {nombre}")

    todos = []
    for un in UN_ORDEN:
        if un not in encontrados:
            print(f"    ⚠ Falta archivo para: {un}")
            continue
        print(f"    ✓ {un} ← {os.path.basename(encontrados[un])}")
        df = procesar_archivo(encontrados[un], un, periodo)
        todos.append(df)

    return pd.concat(todos, ignore_index=True) if todos else pd.DataFrame()

def agrupar_por_trafo(df):
    """Agrupa a nivel de transformador físico."""
    con = df[df['TRAFO'].notna()].copy()
    return con.groupby(['UN_CODIGO','SUBESTACION','TRAFO']).agg(
        MVA_OA          = ('MVA_OA','first'),
        MVA_FOA         = ('MVA_FOA','first'),
        DEM_MAX_REG_MES = ('DEM_MAX_REG_MES','first'),
        FP_TRAFO        = ('FP_TRAFO','first'),
        CARG_OA         = ('CARG_OA','first'),
        CARG_FOA        = ('CARG_FOA','first'),
        N_ALIM          = ('NOMBRE_ALIM','count'),
        ALIMENTADORES   = ('NOMBRE_ALIM', lambda x: ' · '.join([str(a) for a in x.dropna()])),
    ).reset_index()

def clas(c):
    if pd.isna(c): return 'SIN'
    p = c * 100
    if p >= 90: return 'CRITICO'
    elif p >= 70: return 'ALERTA'
    else: return 'NORMAL'

def alim_a_dict(df_alim, periodo):
    """Convierte el DataFrame de alimentadoras a lista de dicts para el JSON."""
    out = []
    for _, r in df_alim.iterrows():
        if pd.isna(r.get('TRAFO')) or str(r.get('TRAFO','')) in ('', 'nan', 'None'):
            continue
        nombre = str(r['NOMBRE_ALIM']).strip() if pd.notna(r.get('NOMBRE_ALIM')) else ''
        codigo = str(r['CODIGO_ALIM']).strip() if pd.notna(r.get('CODIGO_ALIM')) else ''
        # Usar nombre si existe, si no usar código como fallback
        alim_label = nombre if nombre and nombre.upper() not in ('NAN', '') else codigo
        dem_max_alim = r.get('DEM_MAX_MW')
        out.append({
            'periodo':    periodo,
            'un':         r['UN_CODIGO'],
            'subestacion':str(r['SUBESTACION']),
            'trafo':      str(r['TRAFO']),
            'alimentador':alim_label,
            'dem_max':    round(float(dem_max_alim), 3) if pd.notna(dem_max_alim) else None,
        })
    return out

def trafos_a_dict(df_t, periodo):
    out = []
    for _, r in df_t.iterrows():
        out.append({
            'periodo':    periodo,
            'un':         r['UN_CODIGO'],
            'subestacion':str(r['SUBESTACION']),
            'trafo':      str(r['TRAFO']),
            'mva_oa':     round(float(r['MVA_OA']),1)          if pd.notna(r['MVA_OA'])          else None,
            'mva_foa':    round(float(r['MVA_FOA']),1)         if pd.notna(r['MVA_FOA'])         else None,
            'dem_max':    round(float(r['DEM_MAX_REG_MES']),2) if pd.notna(r['DEM_MAX_REG_MES']) else None,
            'fp':         round(float(r['FP_TRAFO']),3)        if pd.notna(r['FP_TRAFO'])        else None,
            'carg_oa':    round(float(r['CARG_OA'])*100,1)     if pd.notna(r['CARG_OA'])         else None,
            'carg_foa':   round(float(r['CARG_FOA'])*100,1)    if pd.notna(r['CARG_FOA'])        else None,
            'nivel_oa':   r['NIVEL_OA'],
            'nivel_foa':  r['NIVEL_FOA'],
            'n_alim':     int(r['N_ALIM']),
            'alimentadores': str(r['ALIMENTADORES']),
        })
    return out

def resumen_un(df_t, df_alim, periodo):
    out = []
    for un in UN_ORDEN:
        s  = df_t[df_t['UN_CODIGO']==un]
        sa = df_alim[df_alim['UN_CODIGO']==un]
        if len(s) == 0: continue
        out.append({
            'un': un, 'nombre': UN_NOMBRES[un], 'periodo': periodo,
            'n_subes':     int(sa['SUBESTACION'].nunique()),
            'n_alim':      int(len(sa)),
            'n_trafos':    int(len(s)),
            'critico_oa':  int((s['NIVEL_OA']=='CRITICO').sum()),
            'alerta_oa':   int((s['NIVEL_OA']=='ALERTA').sum()),
            'normal_oa':   int((s['NIVEL_OA']=='NORMAL').sum()),
            'critico_foa': int((s['NIVEL_FOA']=='CRITICO').sum()),
            'alerta_foa':  int((s['NIVEL_FOA']=='ALERTA').sum()),
            'normal_foa':  int((s['NIVEL_FOA']=='NORMAL').sum()),
            'mva_total_oa':  round(float(s['MVA_OA'].sum()),1),
            'mva_total_foa': round(float(s['MVA_FOA'].sum()),1),
            'dem_total':     round(float(s['DEM_MAX_REG_MES'].sum()),2),
            'carg_prom_oa':  round(float(s['CARG_OA'].mean()*100),1),
            'carg_prom_foa': round(float(s['CARG_FOA'].mean()*100),1),
        })
    return out

def totales_mes(df_t, df_alim, periodo):
    return {
        'periodo':     periodo,
        'n_un':        11,
        'n_subes':     int(df_alim.groupby('UN_CODIGO')['SUBESTACION'].nunique().sum()),
        'n_alim':      int(len(df_alim)),
        'n_trafos':    int(len(df_t)),
        'critico_oa':  int((df_t['NIVEL_OA']=='CRITICO').sum()),
        'alerta_oa':   int((df_t['NIVEL_OA']=='ALERTA').sum()),
        'normal_oa':   int((df_t['NIVEL_OA']=='NORMAL').sum()),
        'critico_foa': int((df_t['NIVEL_FOA']=='CRITICO').sum()),
        'alerta_foa':  int((df_t['NIVEL_FOA']=='ALERTA').sum()),
        'normal_foa':  int((df_t['NIVEL_FOA']=='NORMAL').sum()),
        'mva_total_oa':  round(float(df_t['MVA_OA'].sum()),1),
        'mva_total_foa': round(float(df_t['MVA_FOA'].sum()),1),
        'dem_total':     round(float(df_t['DEM_MAX_REG_MES'].sum()),2),
    }

# ─── BÚSQUEDA AUTOMÁTICA DE CARPETAS ──────────────────────────────────────────
def buscar_carpetas_disponibles():
    """
    Busca todas las carpetas de meses disponibles en 2025 y 2026.
    Reconoce cualquier formato: 'Enero 2025', '1. Enero 2025', '01. Enero 2025', etc.
    Devuelve lista de (periodo, carpeta) ordenada cronológicamente.
    """
    NOMBRES_MESES = {
        'ENERO':1,'FEBRERO':2,'MARZO':3,'ABRIL':4,'MAYO':5,'JUNIO':6,
        'JULIO':7,'AGOSTO':8,'SEPTIEMBRE':9,'OCTUBRE':10,'NOVIEMBRE':11,'DICIEMBRE':12
    }

    encontradas = []

    for anio, base in [('2025', BASE_2025), ('2026', BASE_2026)]:
        if not os.path.isdir(base):
            continue
        for carpeta_nombre in os.listdir(base):
            carpeta_path = os.path.join(base, carpeta_nombre)
            if not os.path.isdir(carpeta_path):
                continue
            # Buscar nombre del mes en el nombre de la carpeta
            nombre_upper = carpeta_nombre.upper()
            mes_num = None
            for nombre_mes, num in NOMBRES_MESES.items():
                if nombre_mes in nombre_upper and anio in carpeta_nombre:
                    mes_num = num
                    mes_nombre = nombre_mes.capitalize()
                    break
            if mes_num is None:
                continue
            # Verificar que tiene archivos Excel
            archivos = glob.glob(os.path.join(carpeta_path, '*.xlsx')) + \
                       glob.glob(os.path.join(carpeta_path, '*.xls'))
            if len(archivos) == 0:
                continue
            periodo = f"{anio}-{mes_num:02d}"
            encontradas.append((periodo, carpeta_path, mes_nombre, anio, len(archivos)))

    # Ordenar cronológicamente
    encontradas.sort(key=lambda x: x[0])
    return encontradas


# ─── PROGRAMA PRINCIPAL ────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 60)
    print("PROCESADOR DE CARGABILIDAD CNEL EP")
    print("=" * 60)

    todos_trafos   = []
    todos_alim     = []
    todos_unidades = []
    todos_totales  = []
    periodos_ok    = []

    carpetas = buscar_carpetas_disponibles()

    if not carpetas:
        print("\n❌ No se encontraron carpetas con datos.")
        print(f"   Verifica que existan archivos Excel en:")
        print(f"   {BASE_2025}")
        print(f"   {BASE_2026}")
        input("\nPresiona Enter para salir...")
        exit()

    print(f"\nSe encontraron {len(carpetas)} meses con datos:")
    for periodo, _, mes, anio, n_arch in carpetas:
        print(f"  ✓ {mes} {anio} ({n_arch} archivos)")

    print()

    for periodo, carpeta, mes_nombre, anio, _ in carpetas:
        archivos = glob.glob(os.path.join(carpeta, '*.xlsx')) + \
                   glob.glob(os.path.join(carpeta, '*.xls'))

        print(f"\n📁 {mes_nombre} {anio} ({len(archivos)} archivos)")
        df_alim = procesar_carpeta_mes(carpeta, periodo)

        if len(df_alim) == 0:
            print(f"   ⚠ Sin datos para {mes_nombre} {anio}")
            continue

        df_t = agrupar_por_trafo(df_alim)
        df_t['NIVEL_OA']  = df_t['CARG_OA'].apply(clas)
        df_t['NIVEL_FOA'] = df_t['CARG_FOA'].apply(clas)

        print(f"   → {df_alim['SUBESTACION'].nunique()} subestaciones · "
              f"{len(df_alim)} alimentadores · {len(df_t)} transformadores")

        # Resumen críticos FOA
        crit = (df_t['NIVEL_FOA']=='CRITICO').sum()
        aler = (df_t['NIVEL_FOA']=='ALERTA').sum()
        print(f"   → FOA: {crit} críticos 🔴  {aler} en alerta 🟠")

        todos_trafos   += trafos_a_dict(df_t, periodo)
        todos_alim     += alim_a_dict(df_alim, periodo)
        todos_unidades += resumen_un(df_t, df_alim, periodo)
        todos_totales.append(totales_mes(df_t, df_alim, periodo))
        periodos_ok.append(periodo)

    if not periodos_ok:
        print("\n❌ No se encontraron carpetas con datos. Verifica las rutas.")
        input("\nPresiona Enter para salir...")
        exit()

    # Construir JSON
    datos = {
        'periodos':            periodos_ok,
        'periodos_label':      {p: MESES_LABELS.get(p, p) for p in periodos_ok},
        'periodos_label_corto':{p: MESES_LABELS_CORTO.get(p, p) for p in periodos_ok},
        'un_orden':            UN_ORDEN,
        'un_nombres':          UN_NOMBRES,
        'totales':             todos_totales,
        'unidades':            todos_unidades,
        'trafos':              todos_trafos,
        'alimentadoras':       todos_alim,
    }

    # Generar JSON de datos
    datos_json = json.dumps(datos, ensure_ascii=False)

    # Leer plantilla HTML e inyectar datos
    if not os.path.exists(PLANTILLA_HTML):
        print(f"\n❌ No se encontró plantilla.html en: {PLANTILLA_HTML}")
        print(f"   Asegúrate de que plantilla.html esté en la misma carpeta que este script.")
        input("\nPresiona Enter para salir...")
        exit()

    with open(PLANTILLA_HTML, 'r', encoding='utf-8') as f:
        plantilla = f.read()

    html_final = plantilla.replace('__DATOS__', datos_json)

    # Guardar index.html
    os.makedirs(os.path.dirname(SALIDA_HTML), exist_ok=True)
    with open(SALIDA_HTML, 'w', encoding='utf-8') as f:
        f.write(html_final)

    size_kb = os.path.getsize(SALIDA_HTML) / 1024
    print(f"\n{'='*60}")
    print(f"✅ index.html generado correctamente")
    print(f"   Periodos: {periodos_ok}")
    print(f"   Trafos totales: {len(todos_trafos)}")
    print(f"   Alimentadoras totales: {len(todos_alim)}")
    print(f"   Tamaño: {size_kb:.1f} KB")
    print(f"   Guardado en: {SALIDA_HTML}")
    print(f"{'='*60}")
    print(f"\n➡ Ahora abre GitHub Desktop y haz Commit + Push")
    print(f"  Tu dashboard se actualizará en 2-3 minutos.")

    input("\nPresiona Enter para salir...")
